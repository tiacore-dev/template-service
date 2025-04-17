from io import BytesIO
from docxtpl import DocxTemplate
from loguru import logger
from openpyxl import load_workbook
from app.utils.context_builders import format_date, flatten_context, deep_defaultdict


def preprocess_dates(context: dict) -> dict:
    def format_value(key, val):
        if isinstance(val, (dict, list)):
            return preprocess_dates(val)
        if isinstance(key, str) and "date" in key.lower():
            return format_date(val)
        return val

    if isinstance(context, dict):
        return {k: format_value(k, v) for k, v in context.items()}
    elif isinstance(context, list):
        return [preprocess_dates(item) for item in context]
    return context


async def handle_docs(data: dict, template_bytes: bytes, extention: str):
    try:
        logger.debug(f"[handle_docs] Обработка шаблона: .{extention}")
        prepared_data = preprocess_dates(data)
        context = deep_defaultdict(prepared_data)
        if extention == "docx":
            return generate_docx_from_bytes(template_bytes, context)
        elif extention == "xlsx":
            return generate_excel_from_template(template_bytes, context)
        else:
            logger.error(
                f"[handle_docs] Неподдерживаемое расширение: {extention}")
            return None
    except Exception as e:
        logger.exception(f"[handle_docs] Ошибка при генерации документа: {e}")
        return None


def generate_docx_from_bytes(template_bytes: bytes, context: dict) -> bytes:
    try:
        doc_stream = BytesIO(template_bytes)
        doc = DocxTemplate(doc_stream)

        logger.debug("[generate_docx_from_bytes] Рендер docx шаблона")
        doc.render(context)

        output_stream = BytesIO()
        doc.save(output_stream)
        output_stream.seek(0)
        return output_stream.read()
    except Exception:
        logger.exception(
            "[generate_docx_from_bytes] Ошибка при рендеринге .docx")
        return b""


def generate_excel_from_template(template_bytes: bytes, context: dict) -> bytes:
    try:
        flat_ctx = flatten_context(context)

        input_stream = BytesIO(template_bytes)
        workbook = load_workbook(input_stream)
        sheet = workbook.active

        logger.debug("[generate_excel_from_template] Рендер xlsx шаблона")

        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    for key, value in flat_ctx.items():
                        placeholder = f"{{{{ {key} }}}}"
                        if placeholder in cell.value:
                            old_val = cell.value
                            cell.value = cell.value.replace(
                                placeholder, str(value))
                            logger.debug(
                                f"[xlsx] Заменено: {old_val} -> {cell.value}")

        output_stream = BytesIO()
        workbook.save(output_stream)
        output_stream.seek(0)
        return output_stream.read()
    except Exception:
        logger.exception(
            "[generate_excel_from_template] Ошибка при рендеринге .xlsx")
        return b""
